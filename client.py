import argparse
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import date, datetime

def parse_args():
    parser = argparse.ArgumentParser(description="vehicle data processing.")
    parser.add_argument(
        "-k", "--keys",
        nargs="+",
        required=True,
        help="keys to process, e.g. kurzname info"
    )
    parser.add_argument(
        "-c", "--colored",
        action="store_true",
        default=True,
        help="enable colored output"
    )
    parser.add_argument(
        "--no-colored",
        action="store_false",
        dest="colored",
        help="disable colored output"
    )
    return parser.parse_args()

def read_csv_data():
    df = pd.read_csv("vehicles.csv", sep=";", quotechar='"', escapechar='\\')
    df = df.fillna("")
    return df.to_dict(orient="records")

def fetch_data_from_api(csv_data):
    response = requests.post("http://localhost:8000/vehicles", json=csv_data)
    response.raise_for_status()
    return response.json()

def save_to_excel(data, keys, colored):
    df = pd.DataFrame(data)
    df = df.sort_values("gruppe") if "gruppe" in df.columns else df
    columns = ["rnr"] if "rnr" in df.columns else []
    columns += [k for k in keys if k in df.columns]
    df = df[columns]
    filename = f"vehicles_{date.today().isoformat()}.xlsx"
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active
    col_indices = {col: idx+1 for idx, col in enumerate(df.columns)}

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        # hu colors
        if colored and "hu" in df.columns:
            hu_value = getattr(row, "hu", None)
            if hu_value:
                try:
                    hu_date = datetime.strptime(str(hu_value), "%Y-%m-%d")
                    diff = (date.today() - hu_date.date()).days
                    if diff <= 90:
                        fill = PatternFill(start_color="FF007500", end_color="FF007500", fill_type="solid")
                    elif diff <= 365:
                        fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFb30000", end_color="FFb30000", fill_type="solid")
                    for col in range(1, len(columns)+1):
                        ws.cell(row=row_idx, column=col).fill = fill
                except Exception:
                    pass
        # coloring labelIds
        if "labelIds" in keys and "colorCode" in df.columns:
            label_ids = getattr(row, "labelIds", None)
            color_code = getattr(row, "colorCode", None)
            if label_ids and color_code and str(color_code).strip() and color_code != "nan":
                # chech color code format
                if color_code.startswith("#") and len(color_code) == 7:
                    cell = ws.cell(row=row_idx, column=col_indices["labelIds"])
                    cell.font = Font(color=color_code.replace("#", "FF"))

    wb.save(filename)
    print(f"Excel dosyası oluşturuldu: {filename}")

def main():
    args = parse_args()
    csv_data = read_csv_data()
    data = fetch_data_from_api(csv_data)["vehicles"]
    save_to_excel(data, args.keys, args.colored)

if __name__ == "__main__":
    main()
